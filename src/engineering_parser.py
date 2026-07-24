from openpyxl import load_workbook

SECTION_NAMES = {
    "Technical Data Sheet": "General",
    "OPEATIONAL DATA": "Operational Data",
    "MECHANICAL DATA": "Mechanical Data",
    "MOTOR DATA": "Motor Data",
    "MATERIAL OF CONSTRUCTION": "Material of Construction",
}


def parse_engineering_sheet(file_path):
    """
    Reads the engineering Excel sheet and groups parameters
    according to their sections.
    """

    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    current_section = "General"

    pump_data = {
        current_section: {}
    }

    for row in sheet.iter_rows(values_only=True):

        if len(row) < 3:
            continue

        parameter = row[1]
        value = row[2]

        if parameter is None:
            continue

        parameter = str(parameter).strip()

        if parameter == "":
            continue

        # Check if this row is a section heading
        if parameter in SECTION_NAMES:

            current_section = SECTION_NAMES[parameter]

            if current_section not in pump_data:
                pump_data[current_section] = {}

            continue

        # Store parameter inside current section
        if value is not None:
            pump_data[current_section][parameter] = value

    return pump_data